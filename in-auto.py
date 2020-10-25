from selenium import webdriver         # Webブラウザを自動操作する（python -m pip install selenium)
from selenium.webdriver.common.keys import Keys # webdriverからスクレイピングで使用するキーを使えるようにする。
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
import chromedriver_binary
import os
import sys
import time
import openpyxl
import datetime

now = datetime.datetime.now()

wb = openpyxl.load_workbook('list.xlsx',data_only=True)
login = openpyxl.load_workbook('user.xlsx',data_only=True)

sheet = wb["Sheet1"]
sheet2 = login["sheet2"]

print("自動入庫を開始します。")
print("salesforceへのログイン情報確認")

user = sheet2.cell(row=2,column=2).value
fpass = sheet2.cell(row=3,column=2).value


if not user:
    print('ユーザー名が未入力です')
    sys.exit()
elif not fpass:
    print('パスワードが未入力です')
    sys.exit()
else:
    print("ユーザーID:"+str(user))
    print("パスワード:"+str(fpass))

print('よろしいですか?y/n')
select = input()

if select== "n":
    print('終了します。')
    sys.exit()
else:
    print()

karide = sheet.cell(row=6,column=1).value#←エクセルから読み取り
pn = sheet.cell(row = 9,column = 1).value#←エクセルから読み取り
sn = sheet.cell(row = 11,column = 1).value#←エクセルから読み取り
times = now.strftime('20%y/%m/%d')
qty = sheet.cell(row=52,column = 1).value

if not karide:
    print('仮出番号が未入力です')
    sys.exit()
elif not pn:
    print('P/Nが未入力です')
    sys.exit()
elif not sn:
    print('S/Nが未入力です')
    sys.exit()
else:
    print()

print('以下内容で入庫開始します。')
print('仮出番号:'+karide)
print('P/N:'+pn)
print('数量:'+str(qty))

print('実行しますか?y/n')
select = input()

if select== "n":
    print('終了します。')
    sys.exit()
else:
    print()

time.sleep(3)

#---------------------------------

#options = webdriver.ChromeOptions()
#options.add_argument('--user-data-dir=C:UsersUserNameAppDataLocalGoogleChromeUser Data')
#options.add_argument('--profile-directory=Default')  # この行を省略するとDefaultフォルダが指定されます
userdata_dir = 'UserData'
os.makedirs(userdata_dir, exist_ok=True)

options = webdriver.ChromeOptions()
options.add_argument('--user-data-dir=' + userdata_dir)
driver = webdriver.Chrome(options=options)

driver.get('https://login.salesforce.com/')  # salesforceを開く
input = driver.find_element_by_id('username')#ユーザー入力
input.send_keys(user)
input = driver.find_element_by_id('password')#パスワード入力
input.send_keys(fpass)
input = driver.find_element_by_id('Login').click()

time.sleep(3)

def inauto() :
    driver.get('https://ap3.salesforce.com/a00/e?retURL=%2Fa00%2Fo')
    time.sleep(3)

    reason = driver.find_element_by_name("00N10000003ccSh")
    reason_select = Select(reason)
    reason_select.select_by_value("New")

    time.sleep(3)

    input = driver.find_element_by_id('CF00N10000002YqNJ')#P/N
    input.send_keys(pn)

    input = driver.find_element_by_id('00N10000002YqNO')#QTY
    input.send_keys(1)

    input = driver.find_element_by_id('00N10000002YqNT')#S/N
    input.send_keys(sn)

    input = driver.find_element_by_id('00N10000002YqNs')
    input.send_keys(times)

    input = driver.find_element_by_id('00N5F000006efkq')#NOte2
    input.send_keys(karide)

    list = driver.find_element_by_name("00N10000002YqNn")
    list_select = Select(list)
    list_select.select_by_value("ITSS 東京")

    selector = '#topButtonRow > input:nth-child(1)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)

    selector = '#topButtonRow > input:nth-child(5)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)

    time.sleep(3)

inauto()
c = 12

while qty > 1 :
    sn =sheet.cell(row = c,column = 1).value
    input = driver.find_element_by_id('00N10000002YqNT').clear()
    input = driver.find_element_by_id('00N10000002YqNT')
    input.send_keys(sn)

    qty = qty - 1
    c = c + 1

    selector = '#topButtonRow > input:nth-child(1)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)
    if qty == 1 :
     continue
    selector = '#topButtonRow > input:nth-child(5)'
    element = driver.find_element_by_css_selector(selector)
    driver.execute_script('arguments[0].click();', element)
else :
    print('入庫終了します。')
    driver.quit()
    sys.exit()

sys.exit()
